from django.core.cache import cache
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework import viewsets, permissions
from rest_framework.views import APIView
from django.db.models.functions import TruncDate, TruncMonth, Coalesce
from django.db.models import Q, Sum, Case, When, F, Count, Value, FloatField, DecimalField
from .utils import calculate_performance_score, detect_fraud_flags
from .serializers import MonthlyLeagueSnapshotSerializer
from tax.models import TaxEntry, MonthlyLeagueSnapshot
from users.models import CustomUser
from payments.models import Payment
import csv
from django.http import HttpResponse
from .permissions import IsOversightRole

from .models import PerformanceTarget
import calendar
from datetime import datetime
from django.utils import timezone
from birs_django.utils.date_utils import get_current_period, get_last_month_period, get_current_quarter



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def performance_summary(request):
  month, year = get_current_period()
  try:
    user = request.user
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    use_date_range = bool(from_date and to_date)

    # Get target

    target_obj = PerformanceTarget.objects.filter(
        user=user,
        month=month,
        year=year
    ).order_by("-created_at").first()

    target_amount = target_obj.target_amount if target_obj else 0

    base_qs = TaxEntry.objects.filter(user=user)
    if use_date_range:
        base_qs = base_qs.filter(date_of_remittance__range=[from_date, to_date])
    else:
        base_qs = base_qs.filter(month=month, year=year)

    # POS subtotal
    pos_totals = base_qs.filter(source="POS").aggregate(
        remita=Sum('remita_amount'),
        interswitch=Sum('interswitch_amount'),
        gokollect=Sum('gokollect_amount')
    )

    pos_total = (pos_totals['remita'] or 0) + (pos_totals['interswitch'] or 0) + (pos_totals['gokollect'] or 0)


    # Manual subtotal
    manual_totals = base_qs.filter(source="Manual").aggregate(
        remita=Sum('remita_amount'),
        interswitch=Sum('interswitch_amount'),
        gokollect=Sum('gokollect_amount')
    )

    manual_total = (manual_totals['remita'] or 0) + (manual_totals['interswitch'] or 0) + (manual_totals['gokollect'] or 0)

    grand_total = pos_total + manual_total
    percent_met = round((grand_total / target_amount * 100), 2) if target_amount else 0

    # Monthly chart data
    monthly_data = (
        TaxEntry.objects.filter(user=user)
        .values('year', 'month')
        .annotate(
            pos_total=Sum(
                Case(
                    When(
                        source="POS",
                        then=Coalesce(
                            F("remita_amount"),
                            Value(0, output_field=DecimalField())
                        ) + Coalesce(
                            F("interswitch_amount"),
                            Value(0, output_field=DecimalField())
                        ) + Coalesce(
                            F("gokollect_amount"),
                            Value(0, output_field=DecimalField())
                        )
                    ),
                    default=Value(0, output_field=DecimalField()),
                    output_field=FloatField()
                )
            ),
            manual_total=Sum(
                Case(
                    When(
                        source="Manual",
                        then=Coalesce(
                            F("remita_amount"),
                            Value(0, output_field=DecimalField())
                        ) + Coalesce(
                            F("interswitch_amount"),
                            Value(0, output_field=DecimalField())
                        ) + Coalesce(
                            F("gokollect_amount"),
                            Value(0, output_field=DecimalField())
                        )
                    ),
                    default=Value(0, output_field=DecimalField()),
                    output_field=FloatField()
                )
            )
        )
        .order_by('year', 'month')

    )

    chart_labels = [calendar.month_name[row['month']] for row in monthly_data]
    pos_values = [float(row['pos_total'] or 0) for row in monthly_data]
    manual_values = [float(row['manual_total'] or 0) for row in monthly_data]
    total_values = [pos_values[i] + manual_values[i] for i in range(len(pos_values))]

    
    recent_entries = base_qs.order_by('-date_of_remittance')[:10]
    record_data = [{
        "date_of_remittance": (
            rec.date_of_remittance.strftime("%Y-%m-%d")
            if rec.date_of_remittance
            else None
        ),
        "remita_amount": float(rec.remita_amount or 0),
        "interswitch_amount": float(rec.interswitch_amount or 0),
        "gokollect_amount": float(rec.gokollect_amount or 0)
    } for rec in recent_entries]

    summaries = {
        'target': target_amount,
        'pos_total': pos_total,
        'manual_total': manual_total,
        'grand_total': grand_total,
        'percent_met': percent_met,
        'chart_labels': chart_labels,
        'pos_values': pos_values,
        'manual_values': manual_values,
        'total_values': total_values,
        'records': record_data,
    }
    return Response(summaries)
  except Exception as e:
        print("🔥 SUMMARY ERROR:", str(e))
        return Response({"error": str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_my_entries_csv(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    user = request.user
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    entries = TaxEntry.objects.filter(user=user)

    if from_date and to_date:
        entries = entries.filter(date_of_remittance__range=[from_date, to_date])

    entries = entries.order_by('-date_of_remittance')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Submissions"

    station_name = getattr(user, "area_office", None) or user.username
    period_label = f"{from_date} to {to_date}" if (from_date and to_date) else "Current Month"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"BIRS Revenue Submissions — {station_name}"
    ws["A1"].font = Font(size=14, bold=True, color="052E16")
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Period: {period_label}"
    ws["A2"].font = Font(size=10, italic=True, color="6B7280")

    headers = ["Date", "Remita (NGN)", "Interswitch (NGN)", "GoKollect (NGN)", "Total (NGN)"]
    header_row = 4
    header_fill = PatternFill(start_color="052E16", end_color="052E16", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_idx = header_row + 1
    grand_total = 0.0

    for rec in entries:
        remita = float(rec.remita_amount or 0)
        interswitch = float(rec.interswitch_amount or 0)
        gokollect = float(rec.gokollect_amount or 0)
        total = remita + interswitch + gokollect
        grand_total += total

        date_str = rec.date_of_remittance.strftime("%Y-%m-%d") if rec.date_of_remittance else "N/A"

        row_data = [date_str, remita, interswitch, gokollect, total]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

        row_idx += 1

    if entries.exists():
        total_label_cell = ws.cell(row=row_idx, column=1, value="TOTAL")
        total_label_cell.font = Font(bold=True)
        total_label_cell.border = thin_border
        total_label_cell.alignment = Alignment(horizontal="center")

        for col_idx in range(2, 5):
            ws.cell(row=row_idx, column=col_idx, value="").border = thin_border

        total_cell = ws.cell(row=row_idx, column=5, value=grand_total)
        total_cell.font = Font(bold=True, color="052E16")
        total_cell.number_format = "#,##0.00"
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        total_cell.border = thin_border
    else:
        ws.cell(row=row_idx, column=1, value="No records found for the selected period.")
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)

    column_widths = [16, 18, 18, 18, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"my_submissions_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class MonthlyLeagueSnapshotViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = MonthlyLeagueSnapshotSerializer
    permission_classes = [permissions.IsAdminUser]

    def get_queryset(self):
        queryset = MonthlyLeagueSnapshot.objects.all().order_by('-year', '-month')

        # Get query params
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')

        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)

        return queryset

@api_view(["POST"])
@permission_classes([IsAdminUser])
def generate_monthly_snapshot(request):
    now = timezone.now()

    month = now.month
    year = now.year

    # ❌ Prevent duplicate snapshots
    if MonthlyLeagueSnapshot.objects.filter(month=month, year=year).exists():
        return Response({"error": "Snapshot already exists for this month"}, status=400)

    atos = CustomUser.objects.filter(role="ato")

    snapshot_data = []

    for ato in atos:
        entries = TaxEntry.objects.filter(
            user=ato,
            date_of_remittance__month=month,
            date_of_remittance__year=year
        )

        totals = entries.aggregate(
            remita=Coalesce(
                Sum("remita_amount"),
                Value(0, output_field=DecimalField())
            ),
            interswitch=Coalesce(
                Sum("interswitch_amount"),
                Value(0, output_field=DecimalField())
            ),
            gokollect=Coalesce(
                Sum("gokollect_amount"),
                Value(0, output_field=DecimalField())
            )
        )

        total = (
            totals["remita"] +
            totals["interswitch"] +
            totals["gokollect"]
        )

        target = PerformanceTarget.objects.filter(
            user=ato,
            month=month,
            year=year
        ).first()

        target_amount = target.target_amount if target else 0

        percent = (total / target_amount * 100) if target_amount > 0 else 0

        snapshot_data.append({
            "user_id": ato.id,
            "name": getattr(ato, "full_name", ato.username),
            "collected": float(total),
            "target": float(target_amount),
            "percent": round(percent, 1)
        })

    # Rank them
    snapshot_data = sorted(snapshot_data, key=lambda x: x["percent"], reverse=True)

    MonthlyLeagueSnapshot.objects.create(
        month=month,
        year=year,
        data=snapshot_data
    )

    return Response({"message": "Snapshot generated successfully"})


class PerformanceTrendsView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        year = request.query_params.get('year')
        start_month = request.query_params.get('start_month')
        end_month = request.query_params.get('end_month')

        queryset = MonthlyLeagueSnapshot.objects.all().order_by('year', 'month')

        if year:
            queryset = queryset.filter(year=year)
        if start_month and end_month:
            queryset = queryset.filter(
                Q(year=year, month__gte=start_month, month__lte=end_month)
            )

        labels = []
        avg_percents = []
        grand_totals = []

        for snapshot in queryset:
            percents = [row.get("percent", 0) for row in snapshot.data if row.get("percent") is not None]
            avg_percent = round(sum(percents) / len(percents), 2) if percents else 0
            grand_total = sum([row.get("collected", 0) for row in snapshot.data])

            labels.append(f"{calendar.month_abbr[snapshot.month]} {snapshot.year}")
            avg_percents.append(avg_percent)
            grand_totals.append(grand_total)

        chart_data = {
            "labels": labels,
            "avg_percents": avg_percents,
            "grand_totals": grand_totals
        }

        return Response(chart_data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def league_table(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    month = request.GET.get("month")
    year = request.GET.get("year")

    entries = TaxEntry.objects.all()

    # Apply date filters to revenue
    if from_date and to_date:
        entries = entries.filter(
            date_of_remittance__range=[from_date, to_date]
        )
    elif month and year:
        entries = entries.filter(
            date_of_remittance__month=month,
            date_of_remittance__year=year
        )
    else:
        now = timezone.now()
        month = now.month
        year = now.year
        entries = entries.filter(
            date_of_remittance__month=month,
            date_of_remittance__year=year
        )

    atos = (
        CustomUser.objects
        .filter(role="ato")
    )

    # Aggregate revenue ONCE
    revenue_rows = (
        entries.values("user")
        .annotate(
            remita_total=Coalesce(
                Sum("remita_amount"),
                Value(0, output_field=DecimalField()),
                output_field=DecimalField()
            ),
            interswitch_total=Coalesce(
                Sum("interswitch_amount"),
                Value(0, output_field=DecimalField()),
                output_field=DecimalField()
            ),
            gokollect_total=Coalesce(
                Sum("gokollect_amount"),
                Value(0, output_field=DecimalField()),
                output_field=DecimalField()
            ),
        )
    )

    revenue_map = {
        row["user"]: row
        for row in revenue_rows
    }

    # Target query ONCE
    target_queryset = PerformanceTarget.objects.all()

    if from_date and to_date:
        f_dt = datetime.strptime(from_date, "%Y-%m-%d")
        t_dt = datetime.strptime(to_date, "%Y-%m-%d")

        target_queryset = target_queryset.filter(
            Q(year__gt=f_dt.year, year__lt=t_dt.year) |
            Q(year=f_dt.year, month__gte=f_dt.month) |
            Q(year=t_dt.year, month__lte=t_dt.month)
        )

        target_rows = (
            target_queryset.values("user")
            .annotate(
                target_total=Coalesce(Sum("target_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField())
            )
        )

        target_map = {
            row["user"]: float(row["target_total"])
            for row in target_rows
        }

    else:
        now = timezone.now()
        effective_month = int(month) if month else now.month
        effective_year = int(year) if year else now.year

        target_queryset = target_queryset.filter(
            month=effective_month,
            year=effective_year
        )

        target_map = {
            t.user_id: float(t.target_amount)
            for t in target_queryset
        }

    cache_key = f"league_table_{month}_{year}_{from_date}_{to_date}"

    cached_data = cache.get(cache_key)

    if cached_data:
        return Response(cached_data)
    
    data = []

    for ato in atos:
        revenue = revenue_map.get(ato.id, {})

        remita_total = float(revenue.get("remita_total", 0))
        interswitch_total = float(revenue.get("interswitch_total", 0))
        gokollect_total = float(revenue.get("gokollect_total", 0))

        total_revenue = (
            remita_total +
            interswitch_total +
            gokollect_total
        )

        target_val = target_map.get(ato.id, 0)

        percent = round(
            (total_revenue / target_val * 100), 1
        ) if target_val > 0 else 0

        data.append({
            "user_id": ato.id,
            "username": ato.username,
            "station_name": (
                getattr(ato, "area_office", None)
                or ato.username
            ),
            "target": target_val,
            "remita": remita_total,
            "interswitch": interswitch_total,
            "gokollect": gokollect_total,
            "total": total_revenue,
            "percent": percent,
        })

    ranked = sorted(
    data,
    key=lambda x: x["percent"],
    reverse=True
    )

    print("\n")
    print("=" * 100)
    print("LEAGUE TABLE DEBUG")
    print("=" * 100)

    for row in ranked:
        print(
            row["station_name"],
            "| remita:", row["remita"],
            "| interswitch:", row["interswitch"],
            "| gokollect:", row["gokollect"],
            "| total:", row["total"],
        )

    print("=" * 100)
    print("\n")

    cache.set(cache_key, ranked, timeout=60 * 5)

    return Response(ranked)


@api_view(["GET"])
@permission_classes([IsOversightRole]) # Allow authenticated users
def admin_dashboard(request):
    now = timezone.now()
    month = now.month
    year = now.year

    # Check if today is the first day of the month - if so, reset charts to zero
    is_first_day_of_month = now.day == 1

    try:
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")

        # Use TaxEntry instead of Payment for calculations
        entries = TaxEntry.objects.all()

        if from_date and to_date:
            entries = entries.filter(
                date_of_remittance__range=[from_date, to_date]
            )
        else:
            # Default to current month/year
            entries = entries.filter(
                date_of_remittance__month=month,
                date_of_remittance__year=year
            )

        # If it's the first day of the month, show zero data for current month
        if is_first_day_of_month and not (from_date and to_date):
            # Return zero values for all current month calculations
            grand_total = 0
            remita_total = 0
            interswitch_total = 0
            gokollect_total = 0
            total_target = 0
            ato_stats = []
            ranked = []
            tax_item_totals = []
            pos_total = 0
            manual_total = 0
            chart_data = []
        else:
            # Calculate totals from TaxEntry
            remita_total = entries.aggregate(
                total=Sum('remita_amount')
            )['total'] or 0

            interswitch_total = entries.aggregate(
                total=Sum('interswitch_amount')
            )['total'] or 0

            gokollect_total = entries.aggregate(
                total=Sum('gokollect_amount')
            )['total'] or 0

            grand_total = float(remita_total) + float(interswitch_total) + float(gokollect_total)

            # Get all ATOs
            atos = CustomUser.objects.filter(role="ato")

            target_queryset = PerformanceTarget.objects.all()

            if from_date and to_date:
                f_dt = datetime.strptime(from_date, "%Y-%m-%d")
                t_dt = datetime.strptime(to_date, "%Y-%m-%d")

                target_queryset = target_queryset.filter(
                    Q(year__gt=f_dt.year, year__lt=t_dt.year) |
                    Q(year=f_dt.year, month__gte=f_dt.month) |
                    Q(year=t_dt.year, month__lte=t_dt.month)
                )
            else:
                target_queryset = target_queryset.filter(
                    month=month,
                    year=year
                )

            total_target = float(
                target_queryset.aggregate(
                    total=Coalesce(
                        Sum("target_amount"),
                        Value(0, output_field=DecimalField()),
                        output_field=DecimalField()
                    )
                )["total"]
            )
            ato_revenue = (
                entries.values("user")
                .annotate(
                    remita_total=Coalesce(Sum("remita_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                    interswitch_total=Coalesce(Sum("interswitch_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                    gokollect_total=Coalesce(Sum("gokollect_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                    days_active=Count("date_of_remittance", distinct=True),
                )
            )

            target_queryset = PerformanceTarget.objects.all()

            if from_date and to_date:
                f_dt = datetime.strptime(from_date, "%Y-%m-%d")
                t_dt = datetime.strptime(to_date, "%Y-%m-%d")

                target_rows = (
                    target_queryset.filter(
                        Q(year__gt=f_dt.year, year__lt=t_dt.year) |
                        Q(year=f_dt.year, month__gte=f_dt.month) |
                        Q(year=t_dt.year, month__lte=t_dt.month)
                    )
                    .values("user")
                    .annotate(
                        total_target=Coalesce(
                            Sum("target_amount"),
                            Value(0, output_field=DecimalField()),
                            output_field=DecimalField()
                        )
                    )
                )

                target_map = {
                    row["user"]: float(row["total_target"])
                    for row in target_rows
                }

            else:
                target_map = {
                    t.user_id: float(t.target_amount)
                    for t in PerformanceTarget.objects.filter(
                        month=month,
                        year=year
                    )
                }

            revenue_map = {
                item["user"]: item
                for item in ato_revenue
            }

            ato_stats = []

            for ato in atos:
                revenue = revenue_map.get(ato.id, {})

                total = (
                    float(revenue.get("remita_total", 0)) +
                    float(revenue.get("interswitch_total", 0)) +
                    float(revenue.get("gokollect_total", 0))
                )

                target_amount = target_map.get(ato.id, 0.0)

                percent = round(
                    (total / target_amount * 100), 1
                ) if target_amount > 0 else 0

                days_active = revenue.get("days_active", 0)

                try:
                    score = calculate_performance_score(
                        total,
                        target_amount,
                        days_active
                    )
                    flags = []
                except:
                    score, flags = 0, []

                ato_stats.append({
                    "id": ato.id,
                    "name": getattr(ato, "full_name", ato.username),
                    "username": ato.username,   # add this
                    "station_name": getattr(ato, "station", None)
                        or getattr(ato, "station_name", None)
                        or ato.username,  # add this (best for consistency)
                    "target": target_amount,
                    "total": total,
                    "percent": percent,
                    "score": score,
                    "flags": flags,
                })

            ranked = sorted(ato_stats, key=lambda x: (x["percent"], x["score"], x["total"]), reverse=True)

            # Calculate tax item aggregates
            tax_item_aggregates = (
                entries.values("tax_item")
                .annotate(
                    total_revenue=
                    Coalesce(Sum("remita_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()) +
                    Coalesce(Sum("interswitch_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()) +
                    Coalesce(Sum("gokollect_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField())
                )
                .order_by("-total_revenue")
            )

            # 2. Format it for the frontend
            tax_item_totals = [
                {
                    "item_name": item["tax_item"],
                    "total_revenue": float(item["total_revenue"] or 0)
                }
                for item in tax_item_aggregates
            ]

            chart_data = (
                entries

                

                .filter(date_of_remittance__isnull=False)

                .annotate(trunc_month=TruncMonth("date_of_remittance"))
                .values("trunc_month")
                .annotate(
                    remita=Coalesce(Sum("remita_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                    interswitch=Coalesce(Sum("interswitch_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                    gokollect=Coalesce(Sum("gokollect_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                )
                .order_by("trunc_month")
            )

        # Monthly trend from TaxEntry - only use current-year data by default
        if from_date and to_date:
            trend_entries = TaxEntry.objects.filter(
                date_of_remittance__range=[from_date, to_date]
            )
        else:
            trend_entries = TaxEntry.objects.filter(
                date_of_remittance__year=year,
                date_of_remittance__month__lte=month
            )

        monthly_trend = (
            trend_entries
            .filter(date_of_remittance__isnull=False)
            .annotate(trunc_month=TruncMonth("date_of_remittance"))
            .values("trunc_month")
            .annotate(
                total=
                Coalesce(
                    Sum("remita_amount"),
                    Value(0, output_field=DecimalField())
                ) +
                Coalesce(
                    Sum("interswitch_amount"),
                    Value(0, output_field=DecimalField())
                ) +
                Coalesce(
                    Sum("gokollect_amount"),
                    Value(0, output_field=DecimalField())
                )
            )
            .order_by("trunc_month")
        )
        chart_data = (
            entries
            

            .filter(date_of_remittance__isnull=False)

            .annotate(trunc_month=TruncMonth("date_of_remittance"))
            .values("trunc_month")
            .annotate(
                remita=Coalesce(Sum("remita_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                interswitch=Coalesce(Sum("interswitch_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                gokollect=Coalesce(Sum("gokollect_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
            )
            .order_by("trunc_month")
        )
        pos_total = entries.filter(source="POS").aggregate(
            total=
            Coalesce(Sum('remita_amount'), Value(0, output_field=DecimalField()), output_field=DecimalField()) +
            Coalesce(Sum('interswitch_amount'), Value(0, output_field=DecimalField()), output_field=DecimalField()) +
            Coalesce(Sum('gokollect_amount'), Value(0, output_field=DecimalField()), output_field=DecimalField())
        )['total']

        manual_total = entries.filter(source="Manual").aggregate(
            total=
            Coalesce(Sum('remita_amount'), Value(0, output_field=DecimalField()), output_field=DecimalField()) +
            Coalesce(Sum('interswitch_amount'), Value(0, output_field=DecimalField()), output_field=DecimalField()) +
            Coalesce(Sum('gokollect_amount'), Value(0, output_field=DecimalField()), output_field=DecimalField())
        )['total']


        # If it's the first day of the month and no custom date filter, set chart_data to empty
        if is_first_day_of_month and not (from_date and to_date):
            chart_data = []
        else:
            chart_data = (
                entries

               

                .filter(date_of_remittance__isnull=False)

                .annotate(trunc_month=TruncMonth("date_of_remittance"))
                .values("trunc_month")
                .annotate(
                    remita=Coalesce(Sum("remita_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                    interswitch=Coalesce(Sum("interswitch_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                    gokollect=Coalesce(Sum("gokollect_amount"), Value(0, output_field=DecimalField()), output_field=DecimalField()),
                )
                .order_by("trunc_month")
            )

        return Response({
            "kpis": {
                "total_revenue": grand_total,
                "total_target": total_target,
                "overall_percent": round((grand_total / total_target * 100), 1) if total_target else 0,
                "total_atos": atos.count(),
                "remita_total": float(remita_total),
                "interswitch_total": float(interswitch_total),
                "gokollect_total": float(gokollect_total),
                "grand_total": grand_total,
            },
            "all_officers": ranked,
            "top5": ranked[:5],
            "bottom5": sorted(
                ranked,
                key=lambda x: x["percent"]
            )[:5],
            "risky": [a for a in ranked if len(a["flags"]) > 0],
            "monthly_trend": [
                {
                    "month": m["trunc_month"].strftime("%b %Y") if m.get("trunc_month") else "N/A",
                    "total": float(m["total"] or 0)
                }
                for m in monthly_trend
            ],
            "chart_data": [
                {
                    "month": m["trunc_month"].strftime("%b"),
                    "remita": float(m["remita"]),
                    "interswitch": float(m["interswitch"]),
                    "gokollect": float(m["gokollect"]),
                }
                for m in chart_data
            ],
            
            "tax_item_totals": tax_item_totals,
            

            "source_breakdown": {
                "pos": float(pos_total),
                "manual": float(manual_total),
            },

        })
        

    except Exception as e:
        # 🔥 THIS WILL NOW PRINT TO YOUR TERMINAL
        import traceback
        print("\n" + "!"*60)
        print("ADMIN DASHBOARD ERROR TRACEBACK:")
        print(traceback.format_exc())
        print("!"*60 + "\n")
        return Response({"error": "Server Error", "details": str(e)}, status=500)
    

def export_revenue_csv(request):
    # 1. Create the HttpResponse object with the appropriate CSV header.
    filename = f"revenue_report_{datetime.now().strftime('%Y-%m-%d')}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # 2. Create the CSV writer using the response as the "file"
    writer = csv.writer(response)
    
    # 3. Write the header row
    writer.writerow(['Month', 'Revenue', 'Payment Method'])

    # 4. Write data rows (For now, static data. Later, pull from your Models/Firestore)
    writer.writerow(['April', '5000000', 'Remita'])
    writer.writerow(['April', '2500000', 'Interswitch'])
    writer.writerow(['April', '0', 'GoKollect'])

    return response


